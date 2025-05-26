import os
from datetime import datetime, timedelta

from aiogram.types import FSInputFile
from loguru import logger
from openpyxl import Workbook

from config import PAIR_TABLE_MAP, ADMIN_ID, ADMIN_ID2
from db_pack.db import get_all_id_with_registered_to_status, get_first_message, user_get_any
from bot.handlers.statistic import TradeStatistics


class AdminTradeStatistics(TradeStatistics):
    async def get_statistics_for_admin(self, ref_date: datetime.date) -> dict:
        all_data = await self._get_all_data_for_user()
        day_str = ref_date.strftime("%d.%m.%Y")
        month_str = ref_date.strftime("%m.%Y")

        stats = {
            "day": {"date": day_str, "data": {}, "total": {"count": 0, "profit": 0.0}},
            "month": {"date": month_str, "data": {}, "total": {"count": 0, "profit": 0.0}},
            "all_time": {"data": {}, "total": {"count": 0, "profit": 0.0}}
        }

        for pair, trades in all_data.items():
            day_count, day_profit = self._calculate_statistics_for_period(trades, day_str, "day")
            month_count, month_profit = self._calculate_statistics_for_period(trades, month_str, "month")
            all_count, all_profit = await self.get_all_period(trades)

            stats["day"]["data"][pair] = {"count": day_count, "profit": round(day_profit, 2)}
            stats["month"]["data"][pair] = {"count": month_count, "profit": round(month_profit, 2)}
            stats["all_time"]["data"][pair] = {"count": all_count, "profit": round(all_profit, 2)}

            stats["day"]["total"]["count"] += day_count
            stats["day"]["total"]["profit"] += day_profit
            stats["month"]["total"]["count"] += month_count
            stats["month"]["total"]["profit"] += month_profit
            stats["all_time"]["total"]["count"] += all_count
            stats["all_time"]["total"]["profit"] += all_profit

        return stats

async def get_statistic_for_admin(bot):
    today = datetime.today().date()
    yesterday = today - timedelta(days=1)
    all_users = await get_all_id_with_registered_to_status(today)

    admin_data = {}

    for user in all_users:
        res = await get_first_message(user)
        user_id = res.from_user.id
        try:
            id_ = await user_get_any(user_id, id='id')
            user_name = await user_get_any(user_id, username='username')
            if user_name == "Нет":
                user_name = await user_get_any(user_id, first_name='first_name')
            date_of_registration = await user_get_any(user_id, registered_at='registered_at')
            end_of_subscription = await user_get_any(user_id, registered_to='registered_to')

            stats_instance = AdminTradeStatistics(user_id, PAIR_TABLE_MAP)
            user_stats = await stats_instance.get_statistics_for_admin(yesterday)

            admin_data[user_id] = {
                "id": id_,
                "username": user_name,
                "registered_at": date_of_registration.strftime("%d.%m.%Y"),
                "registered_to": end_of_subscription.strftime("%d.%m.%Y"),
                "orders": user_stats
            }
            logger.info(f"Статистика для пользователя {user_id} сформирована.")
        except Exception as e:
            logger.warning(f"Ошибка при формировании статистики для пользователя {user}: {e}")

    await generate_excel_file(admin_data, yesterday)
    await send_excel_to_admins(bot, yesterday)



async def generate_excel_file(admin_data: dict, today: datetime.date):
    wb = Workbook()
    day_sheet = wb.active
    day_sheet.title = "day"
    month_sheet = wb.create_sheet("month")
    all_time_sheet = wb.create_sheet("all_time")

    headers = ["ID", "Username", "Дата регистрации", "Подписка до"]
    for pair in PAIR_TABLE_MAP.keys():
        headers.append(f"{pair} - Сделок")
        headers.append(f"{pair} - Прибыль")
    headers.extend(["Всего Сделок", "Всего Прибыль"])

    for sheet in [day_sheet, month_sheet, all_time_sheet]:
        sheet.append(headers)

    for user_id, info in admin_data.items():
        base_data = [
            info["id"],
            info["username"],
            info["registered_at"],
            info["registered_to"]
        ]

        for period, sheet in zip(["day", "month", "all_time"], [day_sheet, month_sheet, all_time_sheet]):
            stats = info["orders"][period]
            row = base_data.copy()

            for pair in PAIR_TABLE_MAP.keys():
                pair_stats = stats["data"].get(pair, {"count": 0, "profit": 0.0})
                row.append(pair_stats["count"])
                row.append(pair_stats["profit"])

            row.extend([stats["total"]["count"], round(stats["total"]["profit"], 2)])
            sheet.append(row)
    for sheet in [day_sheet, month_sheet, all_time_sheet]:
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (AttributeError, TypeError):
                    logger.warning(f"Ошибка при обработке ширины столбца {col_letter}")
            sheet.column_dimensions[col_letter].width = max_length + 2

    filename = f"admin_stats_{today.strftime('%d_%m_%Y')}.xlsx"
    wb.save(filename)
    logger.info(f"Excel файл сохранен: {filename}")



async def send_excel_to_admins(bot, today: datetime.date):
    admin_ids = [ADMIN_ID, ADMIN_ID2]

    filename = f"admin_stats_{today.strftime('%d_%m_%Y')}.xlsx"

    if not os.path.exists(filename):
        logger.error(f"Excel файл {filename} не найден!")
        return

    for admin in admin_ids:
        if admin:
            try:
                file_to_send = FSInputFile(filename)
                await bot.send_document(
                    chat_id=admin,
                    document=file_to_send,
                    caption="Статистика 📊",
                    disable_notification=True
                )
                logger.info(f"Excel файл успешно отправлен админу {admin}")
            except Exception as e:
                logger.error(f"Ошибка при отправке файла админу {admin}: {e}")

    try:
        os.remove(filename)
        logger.info(f"Excel файл {filename} удален после отправки")
    except Exception as e:
        logger.error(f"Ошибка при удалении файла {filename}: {e}")
